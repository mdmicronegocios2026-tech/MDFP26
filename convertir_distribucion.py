import openpyxl
import os

INPUT_FILE = r'C:\Users\neider.cogollo.b\Downloads\XD\Distribución Formadores (1).xlsx'
OUTPUT_FILE = r'C:\Users\neider.cogollo.b\Downloads\XD\Asignaciones_Convertidas.xlsx'

# Evaluator names from header row 5 (columns C-H, 1-based indices 3-8)
EVALUATOR_MAP = {
    3: 'Carlos González',
    4: 'Erika Velosa',
    5: 'Paola Rodríguez',
    6: 'Jimy López',
    7: 'Jairo Izquierdo',
    8: 'Estefanía Rivera'
}

wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)
ws = wb[wb.sheetnames[0]]

students = []
warnings = []
errors = []

for row_idx in range(6, ws.max_row + 1):
    cedula = ws.cell(row=row_idx, column=9).value
    nombre = ws.cell(row=row_idx, column=10).value
    correo = ws.cell(row=row_idx, column=11).value

    if not cedula or not nombre:
        errors.append(f'Row {row_idx}: Missing cedula or name')
        continue

    evaluators = []
    for col_idx, eval_name in EVALUATOR_MAP.items():
        mark = ws.cell(row=row_idx, column=col_idx).value
        if mark and str(mark).strip().lower() in ['evaluador', 'evaluadora']:
            evaluators.append(eval_name)

    if len(evaluators) == 0:
        errors.append(f'Row {row_idx}: {nombre} - No evaluators assigned')
        continue
    elif len(evaluators) > 2:
        warnings.append(f'Row {row_idx}: {nombre} - {len(evaluators)} evaluators, using first 2')

    formador1 = evaluators[0] if len(evaluators) >= 1 else ''
    formador2 = evaluators[1] if len(evaluators) >= 2 else ''

    students.append({
        'cedula': cedula,
        'nombre': nombre,
        'formador1': formador1,
        'formador2': formador2,
        'correo': correo
    })

# Create output workbook
out_wb = openpyxl.Workbook()
out_ws = out_wb.active
out_ws.title = 'Asignaciones'

# Header
out_ws.append(['CEDULA', 'NOMBRE', 'FORMADOR1', 'FORMADOR2', 'CORREO'])

# Data
for s in students:
    out_ws.append([s['cedula'], s['nombre'], s['formador1'], s['formador2'], s['correo']])

out_wb.save(OUTPUT_FILE)

print(f'Total students converted: {len(students)}')
print(f'Warnings: {len(warnings)}')
print(f'Errors: {len(errors)}')
print()
if warnings:
    print('WARNINGS:')
    for w in warnings:
        print(f'  {w}')
    print()
if errors:
    print('ERRORS:')
    for e in errors:
        print(f'  {e}')
    print()
print(f'Output saved to: {OUTPUT_FILE}')

# Summary by evaluator
print('\nEvaluator distribution:')
eval_count = {}
for s in students:
    for e in [s['formador1'], s['formador2']]:
        if e:
            eval_count[e] = eval_count.get(e, 0) + 1
for e, c in sorted(eval_count.items()):
    print(f'  {e}: {c} assignments')
